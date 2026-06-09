"""
DMM Ops Report Generator
========================
Automates the 5-step procedure for preparing the DMM Ops Report.

Input files (expected in the same directory):
  1. EAID AT 7 May for Ops team.xlsx
  2. DMM Ops Report_1-29 April (as on 30 April).xlsx
  3. MTD Revenue_by_Resource_11-05-2026.xlsx

Output:
  DMM Ops Report_UPDATED.xlsx
"""

import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import os
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

EAID_FILE = os.path.join(SCRIPT_DIR, "EAID AT 7 May for Ops team.xlsx")
DMM_OPS_FILE = os.path.join(SCRIPT_DIR, "DMM Ops Report_1-29 April (as on 30 April).xlsx")
RBR_FILE = os.path.join(SCRIPT_DIR, "MTD Revenue_by_Resource_11-05-2026.xlsx")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "DMM Ops Report_UPDATED.xlsx")

# EAID column indices (0-based, headers on row 3, data from row 4)
EAID_HEADER_ROW = 3
EAID_DATA_START_ROW = 4
EAID_COL_EMP_ID = 0       # A: Emp ID
EAID_COL_EMP_NAME = 1     # B: Emp Name
EAID_COL_EMP_EMAIL = 2    # C: Emp Email
EAID_COL_DOJ = 3          # D: DOJ
EAID_COL_LOCATION = 4     # E: Location
EAID_COL_L4 = 8           # I: L4
EAID_COL_DESIGNATION = 10 # K: Designation
EAID_COL_CATEGORY = 16    # Q: Category

DMM_L4_VALUE = "DATA MODERNIZATION & MIGRATION"

# ADMM column indices (0-based)
ADMM_COL_EMP_ID = 0
ADMM_COL_EMP_NAME = 1
ADMM_COL_EMAIL = 2
ADMM_COL_L4 = 3
ADMM_COL_RM_NAME = 4
ADMM_COL_DOJ = 5
ADMM_COL_OFFICE_LOC = 6
ADMM_COL_DESIGNATION = 7
ADMM_COL_CATEGORY = 8

# RBR column index for Employee sub-function
RBR_COL_EMP_SUBFUNCTION = 34  # AI: Employee sub-function
RBR_FILTER_VALUE = "ENG"

# CN = Excel column letter CN = the 92nd column (0-based index 91)
# In RBR this is "FIG Code Description"; in Data tab it's also column 91
CN_COL_INDEX = 91
# We copy columns 0..91 (inclusive) = 92 columns


def read_eaid_dmm_rows():
    """Step 1: Read EAID and filter for DMM (L4 = DATA MODERNIZATION & MIGRATION)."""
    print("[Step 1] Reading EAID file and filtering for DMM...")
    wb = openpyxl.load_workbook(EAID_FILE, read_only=True, data_only=True)
    ws = wb["Base Data"]

    dmm_rows = []
    total = 0
    for row in ws.iter_rows(min_row=EAID_DATA_START_ROW, values_only=True):
        if row[EAID_COL_EMP_ID] is None:
            continue
        total += 1
        l4_val = str(row[EAID_COL_L4]).upper().strip() if row[EAID_COL_L4] else ""
        if l4_val == DMM_L4_VALUE:
            dmm_rows.append(row)

    wb.close()
    print(f"  Total EAID rows scanned: {total}")
    print(f"  DMM rows found: {len(dmm_rows)}")
    return dmm_rows


def read_rbr_eng_rows():
    """Step 5a: Read RBR and filter for ENG in Employee sub-function."""
    print("[Step 5a] Reading RBR file and filtering for ENG...")
    wb = openpyxl.load_workbook(RBR_FILE, read_only=True, data_only=True)
    ws = wb["Export"]

    # Read headers (row 1) — columns A through CN (indices 0..91)
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row[: CN_COL_INDEX + 1])

    eng_rows = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        total += 1
        subfunction = str(row[RBR_COL_EMP_SUBFUNCTION]).upper().strip() if row[RBR_COL_EMP_SUBFUNCTION] else ""
        if subfunction == RBR_FILTER_VALUE:
            # Take only columns A through CN
            eng_rows.append(list(row[: CN_COL_INDEX + 1]))

    wb.close()
    print(f"  Total RBR rows scanned: {total}")
    print(f"  ENG rows found: {len(eng_rows)}")
    return headers, eng_rows


def rebuild_admm(ws_admm, dmm_rows):
    """Step 2 & 3: Clear ADMM and repopulate from EAID DMM rows."""
    print("[Step 3] Rebuilding ADMM sheet...")

    # Preserve header row (row 1)
    # Clear all data rows (row 2 onwards)
    max_row = ws_admm.max_row
    if max_row > 1:
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, 10):  # 9 columns
                ws_admm.cell(row=row_idx, column=col_idx).value = None

    # Write DMM rows
    for i, eaid_row in enumerate(dmm_rows):
        out_row = i + 2  # row 2 onwards

        ws_admm.cell(row=out_row, column=ADMM_COL_EMP_ID + 1).value = eaid_row[EAID_COL_EMP_ID]
        ws_admm.cell(row=out_row, column=ADMM_COL_EMP_NAME + 1).value = eaid_row[EAID_COL_EMP_NAME]
        ws_admm.cell(row=out_row, column=ADMM_COL_EMAIL + 1).value = eaid_row[EAID_COL_EMP_EMAIL]
        ws_admm.cell(row=out_row, column=ADMM_COL_L4 + 1).value = "DMM"
        ws_admm.cell(row=out_row, column=ADMM_COL_RM_NAME + 1).value = None  # RM Name left blank
        ws_admm.cell(row=out_row, column=ADMM_COL_DOJ + 1).value = eaid_row[EAID_COL_DOJ]
        ws_admm.cell(row=out_row, column=ADMM_COL_OFFICE_LOC + 1).value = eaid_row[EAID_COL_LOCATION]
        ws_admm.cell(row=out_row, column=ADMM_COL_DESIGNATION + 1).value = eaid_row[EAID_COL_DESIGNATION]
        ws_admm.cell(row=out_row, column=ADMM_COL_CATEGORY + 1).value = eaid_row[EAID_COL_CATEGORY]

    print(f"  Wrote {len(dmm_rows)} rows to ADMM sheet")


def clean_and_paste_data_tab(ws_data, rbr_headers, rbr_eng_rows):
    """Step 4 & 5: Clean Data tab (keep last 2 cols), paste RBR ENG data."""
    print("[Step 4] Cleaning Data tab...")

    # The Data tab has 94 columns (indices 0-93)
    # Last 2 columns: Practitioner L4 (col 93, CO) and Engg Participating Partner Function (col 94, CP)
    # "Delete columns till Client Name" = delete columns A through BD (indices 0-55)
    # But actually we need to restructure: final layout = RBR columns (A-BD) + last 2 original columns

    # Read the last 2 columns' headers
    last_col_1_header = ws_data.cell(row=1, column=93).value  # Practitioner L4
    last_col_2_header = ws_data.cell(row=1, column=94).value  # Engg Participating Partner Function
    print(f"  Keeping last 2 columns: '{last_col_1_header}', '{last_col_2_header}'")

    # Read existing last-2-column data to preserve if needed (for reference)
    # Per the plan, we leave them blank for new rows

    # Now clear the entire sheet
    max_row = ws_data.max_row
    max_col = ws_data.max_column
    print(f"  Original Data tab: {max_row} rows x {max_col} columns")

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws_data.cell(row=row_idx, column=col_idx).value = None

    # Write RBR headers (columns A through Client Name = 56 columns)
    num_rbr_cols = len(rbr_headers)
    for j, header in enumerate(rbr_headers):
        ws_data.cell(row=1, column=j + 1).value = header

    # Write the last 2 column headers after RBR columns
    ws_data.cell(row=1, column=num_rbr_cols + 1).value = last_col_1_header
    ws_data.cell(row=1, column=num_rbr_cols + 2).value = last_col_2_header

    print(f"[Step 5] Pasting {len(rbr_eng_rows)} ENG rows into Data tab...")

    # Write RBR ENG data rows
    for i, rbr_row in enumerate(rbr_eng_rows):
        out_row = i + 2  # row 2 onwards
        for j, val in enumerate(rbr_row):
            ws_data.cell(row=out_row, column=j + 1).value = val
        # Last 2 columns left blank for new rows

    final_cols = num_rbr_cols + 2
    final_rows = len(rbr_eng_rows) + 1  # +1 for header
    print(f"  Final Data tab: {final_rows} rows x {final_cols} columns")


def main():
    print("=" * 60)
    print("DMM Ops Report Generator")
    print("=" * 60)
    print()

    # Verify all input files exist
    for f in [EAID_FILE, DMM_OPS_FILE, RBR_FILE]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)
    print("All input files found.\n")

    # Step 1: Read EAID and filter for DMM
    dmm_rows = read_eaid_dmm_rows()
    print()

    # Step 5a: Read RBR and filter for ENG (do this before opening DMM for writing)
    rbr_headers, rbr_eng_rows = read_rbr_eng_rows()
    print()

    # Open DMM Ops Report for editing
    print("[Step 2] Opening DMM Ops Report for editing...")
    wb_dmm = openpyxl.load_workbook(DMM_OPS_FILE)
    print(f"  Sheets: {wb_dmm.sheetnames}")
    print()

    # Step 3: Rebuild ADMM sheet
    ws_admm = wb_dmm["ADMM"]
    rebuild_admm(ws_admm, dmm_rows)
    print()

    # Step 4 & 5: Clean Data tab and paste RBR data
    ws_data = wb_dmm["Data"]
    clean_and_paste_data_tab(ws_data, rbr_headers, rbr_eng_rows)
    print()

    # Save output
    print(f"Saving to: {OUTPUT_FILE}")
    wb_dmm.save(OUTPUT_FILE)
    wb_dmm.close()
    print("Done!")

    # Verification summary
    print()
    print("=" * 60)
    print("VERIFICATION SUMMARY")
    print("=" * 60)
    print(f"  ADMM rows written:     {len(dmm_rows)}")
    print(f"  Data tab rows written: {len(rbr_eng_rows)}")
    print(f"  Data tab columns:      {len(rbr_headers) + 2}")
    print(f"  Output file:           {OUTPUT_FILE}")
    print(f"  Output file size:      {os.path.getsize(OUTPUT_FILE) / (1024*1024):.1f} MB")

    # Spot-check
    print()
    print("Spot-check (first 3 ADMM rows):")
    wb_check = openpyxl.load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    ws_check = wb_check["ADMM"]
    for row in ws_check.iter_rows(min_row=1, max_row=4, values_only=True):
        print(f"  {list(row)}")

    print()
    print("Spot-check (Data tab header + first 2 rows):")
    ws_data_check = wb_check["Data"]
    for row in ws_data_check.iter_rows(min_row=1, max_row=3, values_only=True):
        vals = [v for v in row if v is not None]
        print(f"  {vals[:10]}...")

    wb_check.close()
    print()
    print("Report generation complete!")


if __name__ == "__main__":
    main()
