import streamlit as st
import openpyxl
import io
import datetime
from copy import copy

# --- Core Logic Configuration ---
# EAID column indices (0-based, headers on row 3, data from row 4)
EAID_HEADER_ROW = 3
EAID_DATA_START_ROW = 4
EAID_COL_EMP_ID = 0       # A: Emp ID
EAID_COL_EMP_NAME = 1     # B: Emp Name
EAID_COL_EMP_EMAIL = 2    # C: Emp Email
EAID_COL_DOJ = 3          # D: DOJ
EAID_COL_LOCATION = 4     # E: Location
EAID_COL_L4 = 8           # I: L4
EAID_COL_DESIGNATION = 10 # K: Designation
EAID_COL_CATEGORY = 16    # Q: Category

DMM_L4_VALUE = "DATA MODERNIZATION & MIGRATION"

# ADMM column indices (0-based)
ADMM_COL_EMP_ID = 0
ADMM_COL_EMP_NAME = 1
ADMM_COL_EMAIL = 2
ADMM_COL_L4 = 3
ADMM_COL_RM_NAME = 4
ADMM_COL_DOJ = 5
ADMM_COL_OFFICE_LOC = 6
ADMM_COL_DESIGNATION = 7
ADMM_COL_CATEGORY = 8

# RBR column index for Employee sub-function
RBR_COL_EMP_SUBFUNCTION = 34  # AI: Employee sub-function
RBR_FILTER_VALUE = "ENG"

# CN = Excel column letter CN = the 92nd column (0-based index 91)
# In RBR this is "FIG Code Description"; in Data tab it's also column 91
CN_COL_INDEX = 91
# We copy columns 0..91 (inclusive) = 92 columns

# --- Core Logic Functions ---

def read_eaid_dmm_rows(eaid_file):
    """Step 1: Read EAID and filter for DMM (L4 = DATA MODERNIZATION & MIGRATION)."""
    wb = openpyxl.load_workbook(eaid_file, read_only=True, data_only=True)
    ws = wb["Base Data"]

    dmm_rows = []
    total = 0
    for row in ws.iter_rows(min_row=EAID_DATA_START_ROW, values_only=True):
        if row[EAID_COL_EMP_ID] is None:
            continue
        total += 1
        l4_val = str(row[EAID_COL_L4]).upper().strip() if row[EAID_COL_L4] else ""
        if l4_val == DMM_L4_VALUE:
            dmm_rows.append(row)

    wb.close()
    return total, dmm_rows


def read_rbr_eng_rows(rbr_file):
    """Step 5a: Read RBR and filter for ENG in Employee sub-function."""
    wb = openpyxl.load_workbook(rbr_file, read_only=True, data_only=True)
    ws = wb["Export"]

    # Read headers (row 1) — columns A through CN (indices 0..91)
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row[: CN_COL_INDEX + 1])

    eng_rows = []
    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        total += 1
        subfunction = str(row[RBR_COL_EMP_SUBFUNCTION]).upper().strip() if row[RBR_COL_EMP_SUBFUNCTION] else ""
        if subfunction == RBR_FILTER_VALUE:
            # Take only columns A through CN
            eng_rows.append(list(row[: CN_COL_INDEX + 1]))

    wb.close()
    return headers, total, eng_rows


def rebuild_admm(ws_admm, dmm_rows):
    """Step 2 & 3: Clear ADMM and repopulate from EAID DMM rows."""
    # Clear all data rows (row 2 onwards)
    max_row = ws_admm.max_row
    if max_row > 1:
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, 10):  # 9 columns
                ws_admm.cell(row=row_idx, column=col_idx).value = None

    # Write DMM rows
    for i, eaid_row in enumerate(dmm_rows):
        out_row = i + 2  # row 2 onwards

        ws_admm.cell(row=out_row, column=ADMM_COL_EMP_ID + 1).value = eaid_row[EAID_COL_EMP_ID]
        ws_admm.cell(row=out_row, column=ADMM_COL_EMP_NAME + 1).value = eaid_row[EAID_COL_EMP_NAME]
        ws_admm.cell(row=out_row, column=ADMM_COL_EMAIL + 1).value = eaid_row[EAID_COL_EMP_EMAIL]
        ws_admm.cell(row=out_row, column=ADMM_COL_L4 + 1).value = "DMM"
        ws_admm.cell(row=out_row, column=ADMM_COL_RM_NAME + 1).value = None  # RM Name left blank
        ws_admm.cell(row=out_row, column=ADMM_COL_DOJ + 1).value = eaid_row[EAID_COL_DOJ]
        ws_admm.cell(row=out_row, column=ADMM_COL_OFFICE_LOC + 1).value = eaid_row[EAID_COL_LOCATION]
        ws_admm.cell(row=out_row, column=ADMM_COL_DESIGNATION + 1).value = eaid_row[EAID_COL_DESIGNATION]
        ws_admm.cell(row=out_row, column=ADMM_COL_CATEGORY + 1).value = eaid_row[EAID_COL_CATEGORY]


def clean_and_paste_data_tab(ws_data, rbr_headers, rbr_eng_rows):
    """Step 4 & 5: Clean Data tab (keep last 2 cols), paste RBR ENG data."""
    # Read the last 2 columns' headers
    last_col_1_header = ws_data.cell(row=1, column=93).value  # Practitioner L4
    last_col_2_header = ws_data.cell(row=1, column=94).value  # Engg Participating Partner Function

    # Clear the entire sheet
    max_row = ws_data.max_row
    max_col = ws_data.max_column

    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws_data.cell(row=row_idx, column=col_idx).value = None

    # Write RBR headers (columns A through CN = 92 columns)
    num_rbr_cols = len(rbr_headers)
    for j, header in enumerate(rbr_headers):
        ws_data.cell(row=1, column=j + 1).value = header

    # Write the last 2 column headers after RBR columns
    ws_data.cell(row=1, column=num_rbr_cols + 1).value = last_col_1_header
    ws_data.cell(row=1, column=num_rbr_cols + 2).value = last_col_2_header

    # Write RBR ENG data rows
    for i, rbr_row in enumerate(rbr_eng_rows):
        out_row = i + 2  # row 2 onwards
        for j, val in enumerate(rbr_row):
            ws_data.cell(row=out_row, column=j + 1).value = val
        # Last 2 columns left blank for new rows


# --- Streamlit UI ---

st.set_page_config(page_title="DMM Ops Report Generator", layout="wide")

st.title("📊 DMM Ops Report Generator")
st.markdown("""
This application automates the preparation of the DMM Ops Report. 
Please upload the three required Excel files below and click **Generate Report**.
""")

st.divider()

col1, col2, col3 = st.columns(3)

with col1:
    st.subheader("1. EAID File")
    eaid_file = st.file_uploader("Upload EAID List (.xlsx)", type=["xlsx"])

with col2:
    st.subheader("2. DMM Ops Report Template")
    dmm_file = st.file_uploader("Upload DMM Ops Report (.xlsx)", type=["xlsx"])

with col3:
    st.subheader("3. RBR File")
    rbr_file = st.file_uploader("Upload MTD Revenue by Resource (.xlsx)", type=["xlsx"])

st.divider()

if st.button("🚀 Generate Report", use_container_width=True):
    if not (eaid_file and dmm_file and rbr_file):
        st.error("⚠️ Please upload all three required files before generating the report.")
    else:
        try:
            with st.spinner("Step 1/5: Reading EAID file and filtering for DMM..."):
                eaid_total, dmm_rows = read_eaid_dmm_rows(eaid_file)
                st.success(f"✅ EAID processed: {len(dmm_rows)} DMM rows found (out of {eaid_total} total).")

            with st.spinner("Step 2/5: Reading RBR file and filtering for ENG..."):
                rbr_headers, rbr_total, rbr_eng_rows = read_rbr_eng_rows(rbr_file)
                st.success(f"✅ RBR processed: {len(rbr_eng_rows)} ENG rows found (out of {rbr_total} total).")

            with st.spinner("Step 3/5: Opening DMM Ops Report template..."):
                wb_dmm = openpyxl.load_workbook(dmm_file)

            with st.spinner("Step 4/5: Rebuilding ADMM sheet..."):
                ws_admm = wb_dmm["ADMM"]
                rebuild_admm(ws_admm, dmm_rows)
                st.success("✅ ADMM sheet rebuilt successfully.")

            with st.spinner("Step 5/5: Cleaning Data tab and pasting RBR data..."):
                ws_data = wb_dmm["Data"]
                clean_and_paste_data_tab(ws_data, rbr_headers, rbr_eng_rows)
                st.success("✅ Data tab updated successfully.")

            with st.spinner("Saving updated report..."):
                # Save the workbook to a BytesIO object
                output_buffer = io.BytesIO()
                wb_dmm.save(output_buffer)
                output_buffer.seek(0)
            
            st.balloons()
            st.success("🎉 Report generated successfully! You can download it below.")
            
            # Create a download button for the generated file
            current_date = datetime.datetime.now().strftime("%Y-%m-%d")
            st.download_button(
                label="📥 Download Updated DMM Ops Report",
                data=output_buffer,
                file_name=f"DMM_Ops_Report_UPDATED_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            # Display a summary table
            st.subheader("Verification Summary")
            col_a, col_b = st.columns(2)
            col_a.metric("ADMM Rows Written", len(dmm_rows))
            col_b.metric("Data Tab Rows Written", len(rbr_eng_rows))
            
            st.info(f"The updated Data tab contains {len(rbr_headers) + 2} columns.")

        except Exception as e:
            st.error(f"❌ An error occurred during report generation: {str(e)}")
            st.exception(e)

